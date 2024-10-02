import os

from openpyxl import load_workbook, Workbook


def combine_excel_into_one_file(source_folder: str, out_file: str):
    workbook = Workbook()
    sheet = workbook.active

    for root, dirs, filenames in os.walk(source_folder):
        for filename in filenames:
            if filename.endswith('.xlsx'):
                file_path = os.path.join(root, filename)
                try:
                    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
                    for ws in wb.worksheets:
                        for row in ws.iter_rows(values_only=True):
                            sheet.append([filename] + list(row))
                except PermissionError:
                    pass

    workbook.save(out_file)


def read_excel_file(filename, sheet_names: list = None) -> {list}:
    workbook = load_workbook(filename=filename, read_only=True, data_only=True)
    all_data = {}
    if sheet_names is None:
        sheet_names = workbook.get_sheet_names()
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
        all_data[sheet_name] = data
    workbook.close()
    return all_data


def save_excel(file_path, data: list = None, header: list = None):
    workbook = Workbook()
    sheet = workbook.active
    if header:
        sheet.append(header)
    if data:
        for d in data:
            sheet.append(d)
    workbook.save(file_path)
